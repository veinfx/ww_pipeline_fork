import os
import re

from openpyxl import *
from openpyxl.drawing.image import Image
import OpenEXR

from pprint import pprint
import ffmpeg
from ffmpeg import *


class ExcelCreater:

    def __init__(self):

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Shot'

        self._input_path = None
        self.thumbnail_path = None

        self.files_dict = {}

        self.first_file_list = []
        self.last_file_list = []

        self.start_meta = None
        self.last_meta = None

        self.exr_meta_list = []

    @property
    def input_path(self):
        return self._input_path

    @input_path.setter
    def input_path(self, value):
        if value is None or value == "":
            raise ValueError("Input path is missing.")
        self._input_path = value

    def get_all_files(self):
        self.files_dict = {}

        for root, dirs, files in os.walk(self.input_path):
            if files:
                # print(f"ffff==={files}")
                files.sort(key=lambda x: int(re.findall(r'\d+', x)[-1]))
                self.files_dict[root] = files
                sorted(self.files_dict.items(), key=lambda item: item[0], reverse=False)
        if len(self.files_dict.values()) == 0:
            raise Exception("No files found in the directory.")
        # print(f"olol==={self.files_dict}")

        return self.files_dict

    def get_first_and_last_file(self):
        self.first_file_list = []
        self.last_file_list = []

        for root, files in self.files_dict.items():
            if len(files) > 0:
                self.first_file_list.append(root + "/" + files[0])
                self.last_file_list.append(root + "/" + files[-1])
        # print(f"111=={self.first_file_list}, 222=={self.last_file_list}")

        return self.first_file_list, self.last_file_list

    def thumbnail_create(self):
        root_path = self.input_path.split('/production/scan')
        self.thumbnail_path = os.path.join(root_path[0], f'tmp/thumb{root_path[1]}')
        if not os.path.exists(self.thumbnail_path):
            os.makedirs(self.thumbnail_path, exist_ok=True)

        for i, exr in enumerate(self.first_file_list):
            file_name = os.path.splitext(os.path.basename(exr))[0]
            ffmpeg.run(output(input(exr), f'{self.thumbnail_path}/{file_name}.jpg'))

    def get_meta(self):
        # self.origin_data()
        for i, exr in enumerate(self.first_file_list):
            # print(f"bb=={i}=={exr}")
        
            exr_start_file = OpenEXR.InputFile(exr)
            self.start_meta = exr_start_file.header()

            exr_last_file = OpenEXR.InputFile(exr)
            self.last_meta = exr_last_file.header()
            # print(f"333=={self.start_meta}, 444=={self.last_meta}")
            
            file_data = re.match(r"(.*/)([^/]+)\.(\d+)\.(\w+)$", exr)

            # 해상도
            res = re.findall(r'\d+\d+', str(self.start_meta.get("dataWindow")))
            resolutions = list(map(lambda x: str(int(x) + 1), res))

            # 프레임
            frames = re.findall(r'\d+\.\d+|\d+', str(self.start_meta.get("framesPerSecond")))

            self.exr_meta_list.append(
                {
                    "scan_path": file_data.group(1),
                    "scan_name": file_data.group(2),
                    "clip_name": self.start_meta.get("interim.clip.cameraClipName"),
                    "pad": '%0' + str(len(file_data.group(3))) + 'd',
                    "ext": file_data.group(4),
                    "resolutions": ' x '.join(resolutions),
                    "start_frame": int(frames[1]),
                    "and_frame": int(frames[0]),
                    "duration": int(frames[0]) - int(frames[1]) + 1,
                    "timecode_in": self.start_meta.get("arriraw/timeCode"),
                    "timecode_out": self.last_meta.get("arriraw/timeCode"),
                    "framerate":  float(frames[2]),
                    "date": self.start_meta.get("capDate"),
                }
            )

        # print(f"wvwv===={self.exr_meta_list}")

    def thumbnail_data(self):
        # thumbnail
        img_file_list = os.listdir(self.thumbnail_path)

        for i, img_file in enumerate(img_file_list):
            image_path = os.path.join(self.thumbnail_path, img_file)
            image = Image(image_path)
            image.width = 250
            image.height = 150

            col_width = image.width * 50 / 350
            row_height = image.height * 250 / 300

            self.ws.add_image(image, anchor='B' + str(i + 2))
            if i == 0:
                self.ws.column_dimensions['B'].width = col_width
            self.ws.row_dimensions[i + 2].height = row_height
            self.ws.cell(row=i + 2, column=2, value=img_file)

    def execl_form(self):
        header_list = [
            'check', 'thumbnail', 'roll', 'seq_name', 'shot_name', 'version', 'type',
            'scan_path', 'scan_name', 'clip_name', 'pad', 'ext', 'resoultion',
            'start_frame', 'end_frame', 'duration', 'retime_duration', 'retime_percent', 'retime_start_frame',
            'timecode_in', 'timecode_out', 'just_in', 'just_out', 'framerate', 'date', 'clip_tag'
        ]
        for i, title in enumerate(header_list):
            # print(f"titi=={i}")
            self.ws.cell(row=1, column=i + 1, value=title)

    def excel_create(self):
      
        self.execl_form()
        self.thumbnail_data()
        self.get_meta()

        for row, meta in enumerate(self.exr_meta_list, start=2):
            # print(f"coco=={row}==={meta}")

            self.ws.cell(row=row, column=8, value=meta.get("scan_path"))
            self.ws.cell(row=row, column=9, value=meta.get("scan_name"))
            self.ws.cell(row=row, column=10, value=meta.get("clip_name"))
            self.ws.cell(row=row, column=11, value=meta.get("pad"))
            self.ws.cell(row=row, column=12, value=meta.get("ext"))
            self.ws.cell(row=row, column=13, value=meta.get("resolutions"))
            self.ws.cell(row=row, column=14, value=meta.get("start_frame"))
            self.ws.cell(row=row, column=15, value=meta.get("and_frame"))
            self.ws.cell(row=row, column=16, value=meta.get("duration"))

            self.ws.cell(row=row, column=20, value=meta.get("timecode_in"))
            self.ws.cell(row=row, column=21, value=meta.get("timecode_out"))

            self.ws.cell(row=row, column=24, value=meta.get("framerate"))
            self.ws.cell(row=row, column=25, value=meta.get("date"))

        self.excel_save()

    def excel_save(self):
        root_path = self.input_path.split('/production/scan')
        output_path = os.path.join(root_path[0], f'production/excel')
        # print(f"out=={output_path}")
        file_name = os.path.basename(self.input_path)
        # print(f"name==={filename}")
        new_file_name = file_name + '.xlsx'
        save_path = os.path.join(output_path, new_file_name)

        count = 1
        while os.path.exists(save_path):
            new_file_name = f"{file_name}_{count}.xlsx"
            save_path = os.path.join(output_path, new_file_name)
            count += 1

        self.wb.save(save_path)


def main():
    ec = ExcelCreater()

    # setter test info
    ec.input_path = r"/TD/show/hanjin/production/scan/20221017_plate_scan"

    ec.get_all_files()
    ec.get_first_and_last_file()
    ec.thumbnail_create()
    ec.excel_create()


if __name__ == '__main__':
    main()
