import os
from pprint import pprint

import ffmpeg
from openpyxl import Workbook
from openpyxl.drawing.image import Image
# from openpyxl.styles.fonts import Font

# import shotgun_api3

# url = "https://rndtest.shotgrid.autodesk.com/"
# script_name = "dongjin"
# api_key = "n2khFoc&jybfpufdwabdzgxyr"
# sg = shotgun_api3.Shotgun(url, script_name, api_key)

ROOT_DIR = '/TD/show'
project_name = 'hanjin'
project_dir = os.path.join(ROOT_DIR, project_name)
scan_dir = os.path.join(project_dir, 'production/scan')
excel_path = os.path.join(project_dir, 'production/excel')
thumbnail_path = os.path.join(project_dir, 'tmp/thumb')


class CreateExcel:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Shot'

    def get_today_dir(self):
        self.lists = os.listdir(scan_dir)
        self.lists.sort(reverse=True)
        self.dir_name = self.lists[0].split('_')[0]

    def set_header(self):
        header_list = [
            'check', 'thumbnail', 'roll', 'seq_name', 'shot_name', 'version', 'type',
            'scan_path', 'scan_name', 'clip_name', 'pad', 'ext', 'resoultion',
            'start_frame', 'end_frame', 'duration', 'retime_duration', 'retime_percent', 'retime_start_frame',
            'timecode_in', 'timecode_out', 'just_in', 'just_out', 'framerate', 'date', 'clip_tag'
        ]

        for i, title in enumerate(header_list):
            self.ws.cell(row=1, column=i + 1, value=title)

    def get_data(self):
        self.data = []
        root_path = os.path.join(scan_dir, self.lists[0])

        for root, dirs, files in os.walk(root_path):
            temp_list = []
            if len(dirs) > 0:
                for dir in dirs:
                    scan_path = os.path.join(root, dir)

            if len(files) > 0:
                files.sort(reverse=False)
                tmp = files[0].split('.')
                scan_name = tmp[0]
                pad = '%0' + str(len(tmp[1])) + 'd'
                ext = tmp[2]

                temp_list.append(scan_path)
                temp_list.append(scan_name)
                temp_list.append(pad)
                temp_list.append(ext)

            self.data.append(temp_list)

        # for row in self.data:
        #     self.ws.append(row)

        # pprint(len(self.data))
        return self.data

    def insert_thumbnail(self):
        img_dir = thumbnail_path
        img_file_list = os.listdir(img_dir)

        for i, img_file in enumerate(img_file_list):
            image_path = os.path.join(img_dir, img_file)
            image = Image(image_path)
            image.width = 250
            image.height = 150

            col_width = image.width * 50 / 350
            row_height = image.height * 250 / 300

            self.ws.add_image(image, anchor='B' + str(i + 2))
            if i == 0:
                self.ws.column_dimensions['B'].width = col_width  ## 셀 폭은 한 번만 변경
            self.ws.row_dimensions[i + 2].height = row_height  ## 셀 높이 변경
            self.ws.cell(row=i + 2, column=2, value=img_file)  ## 첫 번째 칼럼에 이미지 경로 입력

        # for row in range(len(self.data)):
        #     print(row)
        #     scan_path = self.data[0]
        #     print(scan_path)
        #     scan_name = self.data[1]
        #     pad = self.data[2]
        #     ext = self.data[3]
        #     self.ws.append({'H': scan_path}, {'I': scan_name}, {'k': pad}, {'L': ext})

        # for row in self.deta:
        #     self.ws.append(row=i + 2, column = 8, 9, 11, 12)

    def insert_data(self):
        # pass
        for row in range(len(self.data)):
            scan_path = self.data[0]
            scan_name = self.data[1]
            pad = self.data[2]
            ext = self.data[3]
            self.ws.append({'H': scan_path}, {'I': scan_name}, {'k': pad}, {'L': ext})

    def save_file(self):
        name = self.dir_name + ".xls"
        save_dir_path = os.path.join(excel_path, name)
        self.wb.save(save_dir_path)


if __name__ == "__main__":
    ce = CreateExcel()
    ce.get_today_dir()
    ce.set_header()
    ce.insert_thumbnail()
    ce.get_data()
    ce.insert_data()
    ce.save_file()


