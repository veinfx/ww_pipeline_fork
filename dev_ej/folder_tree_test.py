# import shotgun_api3
#
#
# url = "https://rndtest.shotgrid.autodesk.com/"
# script_name = "dongjin"
# api_key = "n2khFoc&jybfpufdwabdzgxyr"
# sg = shotgun_api3.Shotgun(url, script_name, api_key)

import os
import openpyxl


# from dir_names import *  # Imports the list of folders from dir_names.py file

main_dir = '/home/west/test/show'
ROOT_DIR = 'goguma' #project name(code)
# main_dir_names = ['Gryffindor', 'Slytherin', 'Hufflepuff', 'Ravenclaw']  # Name of the sub-directories


def create_project():
    # shotgrid project code get ?
    os.makedirs(os.path.join(main_dir, ROOT_DIR), exist_ok=True)


def default_dir():
    path = os.path.join(main_dir, ROOT_DIR)
    default_directory = ['assets', 'sequences', 'tmp']
    for d in default_directory:
        os.makedirs(os.path.join(path, d), exist_ok=True)

    os.makedirs(os.path.join(path, 'production'), exist_ok=True)

    p = os.path.join(path, 'production')
    directory = ['scan', 'excel']
    for d in directory:
        os.makedirs(os.path.join(p, d), exist_ok=True)


def get_shot():
    wb = openpyxl.load_workbook('/home/west/test/show/goguma/production/excel/230517.xlsx')
    sheet = wb['Shot']
    for i in range(2, len(sheet['A'])):
        print(sheet.cell(row=i, column=1).value)


def get_seq():
    wb = openpyxl.load_workbook('/home/west/test/show/goguma/production/excel/230517.xlsx')
    sheet = wb['Shot']
    for i in range(2, len(sheet['B'])):
        a = sheet.cell(row=i, column=2).value
        # os.makedirs(os.path.join())


def create_plate_dir():
    path = os.path.join(main_dir, ROOT_DIR)
    default_directory = ['org', 'jpg', 'mp4']
    for d in default_directory:
        os.makedirs(os.path.join(path, d), exist_ok=True)


def find_dir():
    entries = os.listdir(main_dir)
    print(entries)










# def main():
#     # Create directory
#     for i in range(0, len(main_dir)):
#         for j in range(0, len(main_dir[i])):
#             dirName = str(ROOT_DIR) + '/' + str(main_dir_names[i]) + '/' + str(main_dir[i][j])
#
#             try:
#                 # Create target Directory
#                 os.makedirs(dirName)
#                 print("Directory ", dirName, " Created ")
#             except FileExistsError:
#                 print("Directory ", dirName, " already exists")
#
#                 # Create target Directory if don't exist
#             if not os.path.exists(dirName):
#                 os.makedirs(dirName)
#                 print("Directory ", dirName, " Created ")
#             else:
#                 print("Directory ", dirName, " already exists")


if __name__ == '__main__':
    # main()
    # create_project()
    # default_dir()
    # get_shot()
    # get_seq()
    find_dir()

