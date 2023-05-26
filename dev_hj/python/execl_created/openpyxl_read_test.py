from openpyxl import load_workbook

# 엑셀 파일 로드
workbook = load_workbook('/home/west/HJ_root/ihj/production/excel/excel_sample.xlsx')

sheet_names = workbook.sheetnames[0]

print(sheet_names)

# 시트 선택
sheet = workbook[sheet_names]

# # 특정 셀의 값 읽기
# value = sheet['A1'].value
# print(value)
#
# # 특정 범위의 셀 값 읽기
# for row in sheet['A1:C3']:
#     for cell in row:
#         print(cell.value)
#
# # 특정 열 범위의 셀 값 읽기
# for row in sheet.iter_rows(min_row=1, min_col=1, max_row=3, max_col=3):
#     for cell in row:
#         print(cell.value)
#
# # 행 단위로 데이터 읽기
# for row in sheet.iter_rows(min_row=1, max_row=3):
#     values = [cell.value for cell in row]
#     print(values)

# 열 단위로 데이터 읽기
for column in sheet.iter_cols(min_col=2, max_col=3):
    values = [cell.value for cell in column]
    print(values)



column_index = 2

# 열의 값 가져오기
column_values = []
for row in sheet.iter_rows(values_only=True):
    column_values.append(row[column_index - 1])

# 열의 값 출력
for value in column_values:
    print(value)