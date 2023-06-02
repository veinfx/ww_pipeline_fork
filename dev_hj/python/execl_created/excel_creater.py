import os
import re

from openpyxl import *
import OpenEXR


class ExcelCreater:

    def __init__(self):

        self._input_path = None
        self._output_path = None

        self.files_dict = {}

        self.first_file_list = []
        self.last_file_list = []

        self.exr_meta_list = []

    @property
    def input_path(self):
        return self._input_path

    @input_path.setter
    def input_path(self, value):
        if value is None or value == "":
            raise ValueError("Input path is missing.")
        self._input_path = value

    @property
    def output_path(self):
        return self._output_path

    @output_path.setter
    def output_path(self, value):
        self._output_path = value

    def get_all_files(self):
        self.files_dict = {}

        for root, dirs, files in os.walk(self.input_path):
            if files:
                files.sort(key=lambda x: int(re.findall(r'\d+', x)[-1]))
                self.files_dict[root] = files
        if len(self.files_dict.values()) == 0:
            raise Exception("No files found in the directory.")
        print(f"olol==={self.files_dict}")

        return self.files_dict

    def get_first_and_last_file(self):
        self.first_file_list = []
        self.last_file_list = []

        for root, files in self.files_dict.items():
            if len(files) > 0:
                self.first_file_list.append(root + "/" + files[0])
                self.last_file_list.append(root + "/" + files[-1])
        print(f"111=={self.first_file_list}, 222=={self.last_file_list}")

        return self.first_file_list, self.last_file_list

    def exr_metadata(self):
        for i in range(len(self.first_file_list)):
            exr_start_file_path = self.first_file_list[i]
            exr_start_file = OpenEXR.InputFile(exr_start_file_path)
            start_meta = exr_start_file.header()

            exr_last_file_path = self.last_file_list[i]
            exr_last_file = OpenEXR.InputFile(exr_last_file_path)
            last_meta = exr_last_file.header()

            # 해상도
            ress = re.findall(r'\d+\d+', str(start_meta.get("dataWindow")))
            res = list(map(lambda x: str(int(x) + 1), ress))
            resolution = ' x '.join(res)

            # 프레임
            frames = re.findall(r'\d+\.\d+|\d+', str(start_meta.get("framesPerSecond")))
            start_frame = int(frames[1])
            and_frame = int(frames[0])
            framerate = float(frames[2])
            duration = and_frame - start_frame + 1

            self.exr_meta_list.append(
                {
                    "clip_name": start_meta.get("interim.clip.cameraClipName"),
                    "resolutions": resolution,
                    "start_frame": start_frame,
                    "and_frame": and_frame,
                    "duration": duration,
                    "timecode_in": start_meta.get("arriraw/timeCode"),
                    "timecode_out": last_meta.get("arriraw/timeCode"),
                    "just_in": start_frame,
                    "just_out": and_frame,
                    "framerate": framerate,
                    "date": start_meta.get("capDate"),
                    }
                        )
        # print(f"wvwv===={self.exr_meta_list}")

    def excel_create(self):

        # 새로운 엑셀 파일 생성
        wb = Workbook()
        ws = wb.active
        ws.title = 'Shot'

        header_list = [
            'check', 'thumbnail', 'roll', 'seq_name', 'shot_name', 'version', 'type',
            'scan_path', 'scan_name', 'clip_name', 'pad', 'ext', 'resoultion',
            'start_frame', 'end_frame', 'duration', 'retime_duration', 'retime_percent', 'retime_start_frame',
            'timecode_in', 'timecode_out', 'just_in', 'just_out', 'framerate', 'date', 'clip_tag'
        ]

        for i, title in enumerate(header_list):
            # print(f"titi=={i}")
            ws.cell(row=1, column=i + 1, value=title)

        for row, meta in enumerate(self.exr_meta_list, start=2):
            # print(f"coco=={c}==={meta}")

            ws.cell(row=row, column=10, value=meta.get("clip_name"))
            ws.cell(row=row, column=13, value=meta.get("resolutions"))
            ws.cell(row=row, column=14, value=meta.get("start_frame"))
            ws.cell(row=row, column=15, value=meta.get("and_frame"))
            ws.cell(row=row, column=16, value=meta.get("duration"))
            ws.cell(row=row, column=20, value=meta.get("timecode_in"))
            ws.cell(row=row, column=21, value=meta.get("timecode_out"))
            ws.cell(row=row, column=22, value=meta.get("start_frame"))
            ws.cell(row=row, column=23, value=meta.get("and_frame"))
            ws.cell(row=row, column=24, value=meta.get("framerate"))
            ws.cell(row=row, column=25, value=meta.get("date"))

        new_file_name = self.input_path.split("/")[-1] + '.csv'
        save_path = os.path.join(self.output_path, new_file_name)
        count = 1

        while os.path.exists(save_path):
            count += 1
            new_file_name = f"{self.input_path.split('/')[-1]}_{count}.csv"
            save_path = os.path.join(self.output_path, new_file_name)

        wb.save(save_path)


def main():
    ec = ExcelCreater()

    # setter test info
    ec.input_path = r"/home/west/HJ_root/ihj/production/scan/20221018_plate_scan"
    ec.output_path = r"/home/west/HJ_root/ihj/production/excel"

    ec.get_all_files()

    ec.get_first_and_last_file()

    print(f"meta{ec.exr_metadata()}")

    print(f"mack{ec.excel_create()}")


if __name__ == '__main__':
    main()
