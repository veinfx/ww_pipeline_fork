import os
import re

from pprint import *
from datetime import *

from openpyxl import *
import OpenEXR


class ExcelCreated:

    def __init__(self):

        self._input_path = None
        self._output_path = "test"

        self.scan_path = None
        self.scan_name = None

        self.clip_name = None

        self.pad = None
        self.ext = None

        self.resolution = None

        self.start_frame = None
        self.and_frame = None
        self.duration = None

        self.timecode_in = None
        self.timecode_out = None

        self.framerate = None

        self.date = None

        self.files_list = []

    @property
    def input_path(self):
        return self._input_path

    @input_path.setter
    def input_path(self, value):
        if value is None or value == "":
            raise ValueError("Input path is missing.")
        self._input_path = value

    def selected_files(self):
        self.files_list = []
        for file in os.listdir(self.input_path):
            if os.path.isfile(os.path.join(self.input_path, file)):
                self.files_list.append(file)
                self.files_list.sort(key=lambda x: int(re.search(r"\d+", x).group(0)))
        if len(self.files_list) == 0:
            raise Exception("No files found in the directory.")
        return self.files_list

    # def excel_data(self):

    def exr_metadata(self):

        # EXR 파일 경로 설정
        exr_file_path = "/home/west/HJ_root/ihj/production/scan/20221018_plate_scan/001_C140C022_220304_WOFX/C140C022_220304_WOFX.0001014.exr"

        # EXR 파일 열기
        exr_file = OpenEXR.InputFile(exr_file_path)

        # 메타데이터 추출
        header = exr_file.header()
        pprint(header)

        # time_1 = header.get("timeCode")
        # time_2 = str(header.get("capDate"))
        # print(f"time: {time_1} \n \n {time_2} \n ======== \n")

        self.date = header.get("capDate")

        # self.date = str(header.get("capDate"))[2:-1]
        print(f"gogo=== {self.date}")



        # arriraw_1 = str(header.get("arriraw/capDate"))
        # arriraw_2 = str(header.get("arriraw/timeCode"))
        # print(f"arriraw: {arriraw_1} \n \n {arriraw_2} \n ======== \n")

        self.timecode_in = header.get("arriraw/timeCode")
        self.timecode_out = header.get("arriraw/timeCode")

        # self.start_frame = str(header.get("arriraw/timeCode"))[2:-1]
        # self.and_frame = str(header.get("arriraw/timeCode"))[2:-1]
        print(f"tik=== {self.timecode_in}")
        print(f"tok=== {self.timecode_out}")


        # frame_1 = header.get("captureRate")
        # frame_2 = str(header.get("framesPerSecond"))
        # print(f"frame: {frame_1} \n \n {frame_2} \n ======== \n")

        # self.start_frame = frame_2.split("/", 1)

        frame = re.findall(r'\d+\.\d+|\d+', str(header.get("framesPerSecond")))

        self.start_frame = frame[1]
        self.and_frame = frame[0]
        self.framerate = frame[2]
        print(f"jojo=== {self.start_frame}")
        print(f"jojo=== {self.and_frame}")
        print(f"jojo=== {self.framerate}")

        self.clip_name = header.get("interim.clip.cameraClipName")

        # self.clip_name.append(header.get("interim.clip.cameraClipName"))
        print(f"clip=== {self.clip_name}")

    def excel_create(self):

        # 새로운 엑셀 파일 생성
        # workbook = Workbook()
        # sheet = workbook.active

        wb = Workbook()
        ws = wb.active
        ws.title = 'Shot'
        # ws = wb.create_sheet(title="Shot")
        header_list = [
            'check', 'thumbnail', 'roll', 'seq_name', 'shot_name', 'version', 'type',
            'scan_path', 'scan_name', 'clip_name', 'pad', 'ext', 'resoultion',
            'start_frame', 'end_frame', 'duration', 'retime_duration', 'retime_percent', 'retime_start_frame',
            'timecode_in', 'timecode_out', 'just_in', 'just_out', 'framerate', 'date', 'clip_tag'
        ]

        for i, title in enumerate(header_list):
            ws.cell(row=1, column=i + 1, value=title)


        # 파일 이름 시트에 넣기
        # for index, file_name in enumerate(self.files_list, start=2):
        #     print(f"caca==={index}")
        #     ws.cell(row=index, column=9, value=file_name)

        # for index, clip_name in enumerate(self.clip_name, start=2):
        #     print(f"yoyo==={index} // {clip_name}")
        #     ws.cell(row=index, column=10, value=clip_name)

        ws.cell(row=2, column=10, value=self.clip_name)

        ws.cell(row=2, column=14, value=self.start_frame)
        ws.cell(row=2, column=15, value=self.and_frame)

        ws.cell(row=2, column=20, value=self.timecode_in)
        ws.cell(row=2, column=21, value=self.timecode_out)

        ws.cell(row=2, column=24, value=self.framerate)
        ws.cell(row=2, column=25, value=self.date)

        # new_folder = os.path.join(self._output_path, os.path.basename(self._output_path) + '.xlsx')

        save_path = os.path.join("/home/west/HJ_root/ihj/production/excel", "test.xlsx")  # 저장할 경로와 파일명 설정
        wb.save(save_path)


def main():
    ec = ExcelCreated()

    # setter test info
    ec.input_path = r"/home/west/HJ_root/ihj/production/scan/20221018_plate_scan/001_C140C022_220304_WOFX"
    # ec.old_text = ["IHJ", "JOKER"]
    # ec.new_text = ["HJ", "JOK"]

    print(f"하마{ec.selected_files()}")

    print(f"meta{ec.exr_metadata()}")

    print(f"mack{ec.excel_create()}")


if __name__ == '__main__':
    main()