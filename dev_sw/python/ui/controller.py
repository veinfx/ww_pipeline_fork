#!/usr/bin/env python
# :coding: utf-8

import os
import sys

from openpyxl import Workbook

from PySide2.QtWidgets import *
from PySide2.QtCore import *

from model import MyModel
from view import MyView


from dev_sw.python.sg_mp import SgMapping


class MyController:

    def __init__(self, model, view):
        super().__init__()
        # set real path
        now_path = os.path.dirname(os.path.realpath(__file__))
        self.file_path = now_path[:-2] + "/blend_dummy_files"
        self.project_selection_dict = None
        self.seq_selection_dict = None
        self.shot_selection_dict = None

        self.assettype_selection = None
        self.filename_selection = None
        self.filepath_selection = None

        self._project_name = None
        self._asset_info = None
        self._shot_info = None
        self._dir_path = None
        # set real path
        now_path = os.path.dirname(os.path.realpath(__file__))
        self.file_path = now_path[:-6] + 'excel'
        # instance
        self.view = view
        self._dir_path_model = model
        self._dir_path_view = view.excel_path_view
        # set model
        completer = QCompleter(self._dir_path_model)
        self._dir_path_view.setCompleter(completer)

        self.view.save_as_btn.clicked.connect(self.set_excel_path)
        self.view.excel_create_btn.clicked.connect(self.set_excel_save)
        self.view.close_btn.clicked.connect(self.window_close)

        self.get_asset_shot_info_to_url(sys.argv[1])

    def get_asset_shot_info_to_url(self, url):
        """sys.argv 로 받은 샷그리드의 url 정보를 핸들러와 샷그리드 api 로 필요한 정보들을 가져오는 함수이다.
        """
        sa = ShotgunAction(url)
        sg = ShotgunApi()
        self._project_name = sa.project['name']

        entity_type = sa.entity_type
        if entity_type not in ["Asset", "Shot"]:
            raise ValueError("Invaild entity type {}".format(entity_type))

        elif entity_type == "Asset":
            asset_ids = sa.selected_ids
            # asset_ids_filter = sa.selected_ids_filter
            asset_info, _ = sg.get_user_data_info(entity_type, asset_ids)
            self._asset_info = asset_info

        elif entity_type == "Shot":
            shot_ids = sa.selected_ids
            # shot_ids_filter = sa.selected_ids_filter
            _, shot_info = sg.get_user_data_info(entity_type, shot_ids)
            self._shot_info = shot_info

    def set_excel_path(self):
        """
        save as 버튼 클릭시 QFileDialog 를 사용하여 xls파일을 저장할 경로를 설정하여 Qlineedit에 set하는 함수이다.
        """
        self._dir_path_view.clear()
        self._dir_path, check = QFileDialog.getSaveFileName(None, None, self.file_path,  "Xls Files (*.xls);;All Files (*)")
        # self._dir_path_view.setText(os.path.dirname(self._dir_path))
        self._dir_path_view.setText(self._dir_path + '.xls')

    def set_excel_save(self):
        """
        xls save 버튼 클릭시 실행되는 함수이다.
        enumerate() 함수를 사용하여 행 인덱스를 1부터 시작
        """
        workbook = Workbook()
        sheet = workbook.active

        # sheet.title = self._project_name
        # sheet.title = "self._project_name"

        if self._asset_info:
            assets = self._asset_info

            sheet['A1'] = 'Asset Name'
            sheet['B1'] = 'Asset Status'
            sheet['C1'] = 'Asset Type'

            for i, asset in enumerate(assets):
                sheet.cell(row=i+2, column=1, value=asset['code'])
                sheet.cell(row=i+2, column=2, value=asset['sg_status_list'])
                sheet.cell(row=i+2, column=3, value=asset['sg_asset_type'])

        elif self._shot_info:
            shots = self._shot_info

            sheet['A1'] = 'Shot Name'
            sheet['B1'] = 'Shot Status'
            sheet['C1'] = 'Sequence Name'

            for i, shot in enumerate(shots):
                sheet.cell(row=i+2, column=1, value=shot['code'])
                sheet.cell(row=i+2, column=2, value=shot['sg_status_list'])
                sheet.cell(row=i+2, column=3, value=shot['sg_sequence']['name'])

        sheet.column_dimensions['A'].width = 20
        # workbook.save(self._dir_path)
        workbook.save(self._dir_path_view.displayText())
        workbook.close()
        self.view.excel_create_btn_messagebox()

    def window_close(self):
        """
        close 버튼 클릭시 ui 창이 닫힌다.
        """
        self.view.close()


def main():
    QCoreApplication.setAttribute(Qt.AA_ShareOpenGLContexts)
    app = QApplication(sys.argv)
    model = MyModel()
    view = MyView()
    controller = ExcelController(model, view)
    view.show()
    app.exec_()


if __name__ == "__main__":
    main()
