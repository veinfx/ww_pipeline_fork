import os
import pprint

import ffmpeg
from openpyxl import Workbook
from openpyxl.drawing.image import Image
# from openpyxl.styles.fonts import Font

# import shotgun_api3

# url = "https://rndtest.shotgrid.autodesk.com/"
# script_name = "dongjin"
# api_key = "n2khFoc&jybfpufdwabdzgxyr"
# sg = shotgun_api3.Shotgun(url, script_name, api_key)

# project name get ~ scan
root_dir = '/TD/dongjin/projects/gamja/production/scan'
# root_dir = '/TD/show'
# excel_dir = ''


class CreateExcel:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Shot'
        # self.wb.save("C:/Users/jin91/PipelineTD/git/test2.xls")

    def get_today_dir(self):
        list = os.listdir(root_dir)
        # self.
        self.file_name = list[0].split('_')[0]
        print(self.file_name)

    def set_header(self):
        header_list = [
            'check', 'thumbnail', 'roll', 'seq_name', 'shot_name', 'version', 'type',
            'scan_path', 'scan_name', 'clip_name', 'pad', 'ext', 'resoultion',
            'start_frame', 'end_frame', 'duration', 'retime_duration', 'retime_percent', 'retime_start_frame',
            'timecode_in', 'timecode_out', 'just_in', 'just_out', 'framerate', 'date', 'clip_tag'
        ]

        for i, title in enumerate(header_list):
            self.ws.cell(row=1, column=i + 1, value=title)

    def get_col_width_row_height(img_width, img_height):
        col_width = img_width * 50 / 350
        row_height = img_height * 250 / 300
        return col_width, row_height

    def get_scan_path(self):
        pass

    def get_scan_name(self):
        pass

    def get_pad(self):
        pass

    def get_data(self):
        test_dir = os.path.join(root_dir, self.file_name)
        print(test_dir)
        for (root, dirs, files) in os.walk(test_dir):
            print("456456", root)
            print("789789", dirs)
            # print("# root : " + root)
            if len(dirs) > 0:
                for dir_name in dirs:
                    print("dir: " + dir_name)

            if len(files) > 0:
                print("123123", files[0])

        # a = os.listdir(root_dir)[0]
        # directory = os.path.join(root_dir, a)
        # b = os.listdir(directory)
        # c = os.path.abspath(b)
        # print(c)
        # folder_list = list(filter(os.path.isdir, os.listdir(directory)))
        # print(folder_list)

        #result "/TD/dongjin/projects/gamja/production/scan/20221017_plate_scan/001_C140C022_220304_WOFX"

        # get extension
        # test = '/TD/dongjin/projects/gamja/production/scan/20221017_plate_scan/001_C140C022_220304_WOFX/C140C022_220304_WOFX.0001001.exr'
        # ext = os.path.basename(test).split('.')[2]


    def insert_data(self):
        pass

    def insert_thumbnail(self):
        # img_dir = '/home/west/test/images'
        img_dir = 'C:/Users/jin91/PipelineTD/git/images'
        img_file_list = os.listdir(img_dir)

        for i, img_file in enumerate(img_file_list):
            print(i, img_file)
            image_path = os.path.join(img_dir, img_file)
            image = Image(image_path)
            image.height = 150
            image.width = 250
            self.ws.add_image(image, anchor='B' + str(i + 2))
            col_width, row_height = ce.get_col_width_row_height(image.width, image.height)  ## 엑셀 셀 폭 높이 단위
            if i == 0:
                self.ws.column_dimensions['B'].width = col_width  ## 셀 폭은 한 번만 변경
            self.ws.row_dimensions[i + 2].height = row_height  ## 셀 높이 변경
            self.ws.cell(row=i + 2, column=1, value=img_file)  ## 첫 번째 칼럼에 이미지 경로 입력

    def save_file(self):
        # file_name = self.file_name + ".xls"
        # file_name = self.file_name + ".csv"
        # get excel_path add!!!!!!!
        # save_dir_path = os.path.join('/TD/dongjin/projects/gamja/production/excel/', file_name)
        # self.wb.save(save_dir_path)
        self.wb.save('/TD/dongjin/projects/gamja/production/excel/20221017.xls')


if __name__ == "__main__":
    ce = CreateExcel()
    # ce.get_today_dir()
    ce.set_header()
    # ce.get_data()
    ce.save_file()
    # ce.input_thumbnail()

