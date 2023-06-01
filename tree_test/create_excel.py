import os
from pprint import pprint

from openpyxl import Workbook
from openpyxl.drawing.image import Image
# from openpyxl.styles.fonts import Font

from convert_thumb import get_thumbnail

SCAN_PATH = "/TD/show/hanjin/production/scan/20221017_plate_scan"
ROOT_PATH = SCAN_PATH.split('/production/scan')

thumbnail_path = get_thumbnail("/TD/show/hanjin/production/scan/20221017_plate_scan")


class CreateExcel:
    def __init__(self):
        # self.dir_name = os.path.basename(SCAN_PATH).split('_')[0]
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Shot'
        self.input_path = r"C:/Users/jin91/Documents/ShotGrid/show/oksusu/production/scan/20221017_plate_scan"
        self.root_path = self.input_path.split('/production/scan')

    def set_header(self):
        header_list = [
            'check', 'thumbnail', 'roll', 'seq_name', 'shot_name', 'version', 'type',
            'scan_path', 'scan_name', 'clip_name', 'pad', 'ext', 'resoultion',
            'start_frame', 'end_frame', 'duration', 'retime_duration', 'retime_percent', 'retime_start_frame',
            'timecode_in', 'timecode_out', 'just_in', 'just_out', 'framerate', 'date', 'clip_tag'
        ]

        for i, title in enumerate(header_list):
            self.ws.cell(row=1, column=i + 1, value=title)

    def get_data(self):
        # self.data = []
        # for path, dirs, files in os.walk(SCAN_PATH):
        #     temp_list = []
        #     if len(files) == 0:
        #         pass
        #     else:
        #         temp_list.append(path)
        #         files.sort(reverse=False)
        #         tmp = files[0].split('.')
        #         scan_name = tmp[0]
        #         pad = '%0' + str(len(tmp[1])) + 'd'
        #         ext = tmp[2]
        #         temp_list.append(scan_name)
        #         temp_list.append(pad)
        #         temp_list.append(ext)
        #
        #     self.data.append(temp_list)
        # pprint(self.data)
        # return self.data

        self.data = []
        for path, dirs, files in os.walk(SCAN_PATH):
            if len(files) > 0:
                files.sort(reverse=False)
                self.data.append(os.path.join(path, files[0]))
        pprint(self.data)

    def insert_data(self):
        # thumbnail
        img_file_list = os.listdir(thumbnail_path)

        for i, img_file in enumerate(img_file_list):
            image_path = os.path.join(thumbnail_path, img_file)
            image = Image(image_path)
            image.width = 250
            image.height = 150

            col_width = image.width * 50 / 350
            row_height = image.height * 250 / 300

            self.ws.add_image(image, anchor='B' + str(i + 2))
            if i == 0:
                self.ws.column_dimensions['B'].width = col_width
            self.ws.row_dimensions[i + 2].height = row_height
            self.ws.cell(row=i + 2, column=2, value=img_file)

        # other_data
        # for d in self.data:
        #     if len(d) != 0:
        #         for row in len(self.data):
        #             scan_path = d[0]
                    # scan_name = d[1]
                    # pad = d[2]
                    # ext = d[3]
                    # self.ws.append(
                    #     {'H': scan_path}
                        # {'I': scan_name},
                        # {'k': pad}
                        # {'L': ext}
                    # )

    def save_excel_file(self):
        excel_path = os.path.join(ROOT_PATH[0], f'production/excel')
        name = self.dir_name + ".xlsx"
        save_dir_path = os.path.join(excel_path, name)
        self.wb.save(save_dir_path)
    #     # excel_path = os.path.join(self.root_path[0], 'production/excel')
    #     excel_path = 'C:/Users/jin91/Documents/ShotGrid/show/oksusu/production/excel'
    #     # name = self.root_path[1] + '.xlsx'
    #     name = '20221017_plate_scan' + '.xlsx'
    #
    #     save_dir_path = os.path.join(excel_path, name)
    #     print(save_dir_path)
    #
    #     count = 1
    #     print("123")
    #     while os.path.exists(save_dir_path):
    #         print("456")
    #         new_name = f'20221017_plate_scan_{count}.xlsx'
    #         save_dir_path = os.path.join(excel_path, new_name)
    #
    #     self.wb.save(save_dir_path)


if __name__ == "__main__":
    ce = CreateExcel()
    ce.set_header()
    # ce.get_data()
    # ce.insert_data()
    ce.save_excel_file()

