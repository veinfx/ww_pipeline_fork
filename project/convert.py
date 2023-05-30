import os
import csv

# import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image


def get_col_width_row_height(img_width, img_height):
    col_width = img_width * 70 / 504.19
    # row_height = img_height * 225.35 / 298.96
    row_height = img_height * 250 / 298.96
    return col_width, row_height


# wb = Workbook()  ## 워크북 생성
# ws = wb.active  ## 첫 번째 시트
#
# ## 이미지 불러오기
# image_path = '/home/west/test/test.jpg'
#
# image = Image(image_path)
# image.height = 150
# image.width = 250
#
# ws.add_image(image, 'A1')  ## 이미지 삽입
#
# ## 이미지 픽셀을 셀 폭과 높이로 변환
# col_width, row_height = get_col_width_row_height(image.width, image.height)
#
# ws.column_dimensions['A'].width = col_width  ## 셀 폭 변경
# ws.row_dimensions[1].height = row_height  ## 셀 높이 변경
#
# wb.save('/home/west/test/one_image.xlsx')
# wb.close()

wb = Workbook()  ## 워크북 생성
ws = wb.active  ## 첫 번째 시트

img_dir = '/TD/WG_test/project/zombie_city/production/scan/test'
img_file_list = os.listdir(img_dir)  ## 이미지 파일 리스트

for i, img_file in enumerate(img_file_list):
    image_path = os.path.join(img_dir, img_file)  ## 이미지 경로
    image = Image(image_path)  ## 이미지 로드
    image.height = 150
    image.width = 200
    ws.add_image(image, anchor='A' + str(i + 2))  ## 이미지 삽입
    col_width, row_height = get_col_width_row_height(image.width, image.height)  ## 엑셀 셀 폭 높이 단위
    if i == 0:
        ws.column_dimensions['A'].width = col_width  ## 셀 폭은 한 번만 변경
    ws.row_dimensions[i + 2].height = row_height  ## 셀 높이 변경
    ws.cell(row=i + 2, column=1, value=img_file)  ## 첫 번째 칼럼에 이미지 경로 입력

wb.save('/TD/WG_test/project/zombie_city/production/excel/testwg.xlsx')
xlsx_file = '/TD/WG_test/project/zombie_city/production/excel/testwg.xlsx'
csv_file = '/TD/WG_test/project/zombie_city/production/excel/testwg.csv'


wb = load_workbook(filename=xlsx_file)
ws = wb.active

with open(csv_file, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)

    for row in ws.iter_rows(values_only=True):
        writer.writerow(row)

# with open(xlsx_file, 'r', encoding='latin-1', errors='ignore') as f_in, open(csv_file, 'w', newline='',
#                                                                              encoding='utf-8') as f_out:
#     reader = csv.reader(f_in)
#     writer = csv.writer(f_out)
#
#     for row in reader:
#         writer.writerow(row)
