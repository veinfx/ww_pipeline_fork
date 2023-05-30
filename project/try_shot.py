# from openpyxl import load_workbook
# from shotgun_api3 import Shotgun
#
# # ShotGrid 사이트에 연결합니다.
# sg = Shotgun("https://rndtest.shotgrid.autodesk.com", "script_wongyu", "mh3lwvof$rzkUtqndqsfqcckf")
#
# # Excel 파일의 경로를 지정합니다.
# excel_file_path = "/show/excel/hanjin.xlsx"
#
# # sequence와 shot 데이터가 포함된 워크시트의 이름을 지정합니다.
# worksheet_name = "Sheet1"  # 실제 워크시트 이름으로 업데이트하세요.
#
# # 워크북을 로드합니다.
# workbook = load_workbook(excel_file_path)
#
# # 워크시트를 이름으로 선택합니다.
# worksheet = workbook[worksheet_name]
#
# # 열 번호를 지정합니다.
# sequence_name_column = "D"  # sequence 이름이 있는 열 (예: A열)
# shot_name_column = "E"  # shot 코드가 있는 열 (예: B열)
# just_in_column = "V"  # Just-In column (e.g., column V)
# just_out_column = "W"  # Just-Out column (e.g., column W)
#
# sequence_name_column_number = ord(sequence_name_column.upper()) - 64
# shot_name_column_number = ord(shot_name_column.upper()) - 64
# just_in_column_number = ord(just_in_column.upper()) - 64
# just_out_column_number = ord(just_out_column.upper()) - 64
#
# # 헤더를 제외한 각 행을 반복합니다. (첫 번째 행은 헤더입니다)
# for row in worksheet.iter_rows(min_row=2, values_only=True):
#     # Excel 파일에서 sequence 이름과 shot 코드를 가져옵니다.
#     sequence_name = row[sequence_name_column_number - 1]
#     shot_name = row[shot_name_column_number - 1]
#     just_in = row[just_in_column_number - 1]
#     just_out = row[just_out_column_number - 1]
#
#     # ShotGrid에 새로운 sequence를 생성합니다.
#     sequence_data = {
#         "code": sequence_name,
#         "project": {"type": "Project", "id": 299},  # 프로젝트 ID로 업데이트하세요.
#         "sg_just_in": just_in,
#         "sg_just_out": just_out,
#     }
#     sequence = sg.create("Sequence", sequence_data)
#
#     # ShotGrid의 sequence 하위에 새로운 shot을 생성합니다.
#     shot_data = {
#         "code": shot_name,
#         "project": {"type": "Project", "id": 299},  # 프로젝트 ID로 업데이트하세요.
#         "sg_sequence": sequence,
#         "sg_just_in": just_in,
#         "sg_just_out": just_out,
#     }
#     shot = sg.create("Shot", shot_data)
#
#     # 생성된 sequence와 shot 코드를 출력합니다.
#     print("Created Sequence:", sequence["code"])
#     print("Created Shot:", shot["code"])
#
# # Add "just_in" field to the Shot entity
# just_in_field_data = {
#     "entity_type": "Shot",
#     "data_type": "text",
#     "name": "sg_sg_just_in",
# }
#
# just_in_field = sg.create("CustomNonProjectEntity01Field", just_in_field_data)
# print("Created Field:", just_in_field["name"])
#
# # Add "just_out" field to the Sequence entity
# just_out_field_data = {
#     "entity_type": "Sequence",
#     "data_type": "text",
#     "name": "sg_sg_just_out",
# }
#
# just_out_field = sg.create("CustomNonProjectEntity01Field", just_out_field_data)
# print("Created Field:", just_out_field["name"])

import openpyxl
import shotgun_api3
from pprint import pprint

SERVER_PATH = "https://rndtest.shotgrid.autodesk.com"
SCRIPT_NAME = "script_wongyu"
SCRIPT_KEY = "mh3lwvof$rzkUtqndqsfqcckf"
sg = shotgun_api3.Shotgun(SERVER_PATH, SCRIPT_NAME, SCRIPT_KEY)


class xlsx:
    def __init__(self):
        self.just_out = []
        self.resolution = []
        self.ext = []
        self.start_frame = []
        self.end_frame = []
        self.duration = []
        self.timecode_in = []
        self.timecode_out = []
        self.framerate = []
        self.date = []
        self.just_in = []
        self.xlsx_data = []
        self.scan_path = []
        self.type = []

    def data_info(self):
        workbook = openpyxl.load_workbook("/TD/show/excel/hanjin.xlsx")
        worksheet = workbook.active
        header = [cell.value for cell in worksheet[1]]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            self.xlsx_data.append(dict(zip(header, row)))
        pprint(self.xlsx_data)

    def sequence_upload(self):
        existing_sequence_codes = set()

        # Get the existing sequence codes from Shotgun
        existing_sequences = sg.find('Sequence', [['project', 'is', {'type': 'Project', 'id': 130}]], ['code'])
        for sequence in existing_sequences:
            existing_sequence_codes.add(sequence['code'])

        for data in self.xlsx_data:
            sequence_code = data['seq_name']

            if sequence_code not in existing_sequence_codes:
                sequence_data = {
                    'code': sequence_code,
                    'project': {'type': 'Project', 'id': 130}  # Replace with the actual project ID
                }
                sg.create('Sequence', sequence_data)
                existing_sequence_codes.add(sequence_code)
            else:
                print(f"Sequence '{sequence_code}' already exists. Skipping creation.")

    def shot_upload(self):
        existing_shot_codes = set()

        # Get the existing shot codes from Shotgun
        existing_shots = sg.find('Shot', [['project', 'is', {'type': 'Project', 'id': 130}]], ['code'])
        for shot in existing_shots:
            existing_shot_codes.add(shot['code'])

        for data in self.xlsx_data:
            sequence_code = data['seq_name']
            shot_code = data['shot_name']
            self.scan_path = data['scan_path']
            self.type = data['type']
            self.just_in = str(data['just_in'])
            self.just_out = str(data['just_out'])
            self.resolution = str(data['resolution'])
            self.ext = data['ext']
            self.start_frame = str(data['start_frame'])
            self.end_frame = str(data['end_frame'])
            self.duration = str(data['duration'])
            self.timecode_in = str(data['timecode_in'])
            self.timecode_out = str(data['timecode_out'])
            self.framerate = str(data['framerate'])
            self.date = str(data['date'])

            if shot_code not in existing_shot_codes:
                sequence_filters = [['code', 'is', sequence_code]]
                sequence_entity = sg.find_one('Sequence', sequence_filters)

                shot_data = {
                    'code': shot_code,
                    'sg_sequence': sequence_entity,
                    'sg_scan_path': self.scan_path,
                    'sg_just_in': self.just_in,
                    'sg_just_out': self.just_out,
                    'sg_shot_type': self.type,
                    'sg_resolution': self.resolution,
                    'sg_ext': self.ext,
                    'sg_start_frame': self.start_frame,
                    'sg_end_frame': self.end_frame,
                    'sg_duration': self.duration,
                    'sg_timecode_in': self.timecode_in,
                    'sg_timecode_out': self.timecode_out,
                    'sg_framerate': self.framerate,
                    'sg_date': self.date,

                    'project': {'type': 'Project', 'id': 130}  # Replace with the actual project ID
                }
                sg.create('Shot', shot_data)
                existing_shot_codes.add(shot_code)
            else:
                print(f"Shot '{shot_code}' already exists. Skipping creation.")


if __name__ == '__main__':
    xx = xlsx()
    xx.data_info()
    xx.sequence_upload()
    xx.shot_upload()
